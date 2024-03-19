import json
import os
import re
from collections import defaultdict
from io import BytesIO
from itertools import chain, count
from pathlib import Path
from typing import Any, Iterable, Literal, Sequence, TypeAlias

import numpy as np
import pandas as pd
from dotenv import load_dotenv
from httpx import get
from openpyxl import load_workbook
from openpyxl.cell import Cell, MergedCell
from openpyxl.worksheet.worksheet import Worksheet

load_dotenv()

ROOT = Path(__file__).parent

STUDENTS_SCHEDULE_URL = os.getenv("STUDENTS_SCHEDULE_URL")
TEACHERS_SCHEDULE_URL = os.getenv("TEACHERS_SCHEDULE_URL")

assert STUDENTS_SCHEDULE_URL, "STUDENTS_SCHEDULE_URL env variable is not set"
assert TEACHERS_SCHEDULE_URL, "TEACHERS_SCHEDULE_URL env variable is not set"

EMPTY_CELLS = ["---", "-x-", "", np.nan]

EVENT_START_TIMES = [
    "08:30",
    "10:20",
    "12:10",
    "14:30",
    "16:20",
]
EVENT_END_TIMES = [
    "10:05",
    "11:55",
    "13:45",
    "16:05",
    "17:35",
]
DAYS = [
    "Понеділок",
    "Вівторок",
    "Середа",
    "Четвер",
    "Пятниця",
]

CellType: TypeAlias = Literal[
    "single",
    "vertical",
    "horizontal",
    "multiple",
]


def fetch_schedule_io(url: str) -> BytesIO:
    response = get(url, follow_redirects=True)
    response.raise_for_status()

    return BytesIO(response.content)


def get_merge_cell_start(sheet: Worksheet, merged: MergedCell) -> Cell | None:
    for rng in sheet.merged_cells.ranges:
        if merged.coordinate in rng:
            return rng.start_cell

    return None


def iter_students_groups(sheet: Worksheet, offset: int) -> Iterable[tuple[int, str]]:
    empty = 0

    for idx in count(5):
        if group := sheet.cell(row=offset, column=idx).value:
            empty = 0
            yield idx, group
        else:
            empty += 1

        if empty > 1:
            break


def iter_students_days_idx(offset: int) -> Iterable[tuple[int, str, str]]:
    start = offset
    for day in DAYS:
        for t in EVENT_START_TIMES:
            yield start, day, t
            start += 2

        start += 1


def read_group_sheet(sheet: Worksheet, col_idx: int, row_offset: int) -> Iterable[tuple[str, int]]:
    def _get_cell(row: int, col: int) -> Cell:
        cell = sheet.cell(row=row, column=col)

        if isinstance(cell, MergedCell):
            return get_merge_cell_start(sheet, cell)

        return cell

    for idx, (row_idx, *_) in enumerate(iter_students_days_idx(row_offset)):
        idx *= 2

        if nominator := _get_cell(row_idx, col_idx).value:
            yield nominator, idx
        if denominator := _get_cell(row_idx + 1, col_idx).value:
            yield denominator, idx + 1


def get_schedule_df(url: str) -> pd.DataFrame:
    with fetch_schedule_io(url) as io:
        workbook = load_workbook(io)

    df = pd.DataFrame.from_records(
        [{"day": day, "time": f"{time}_{event_tp}"} for day in DAYS for time in EVENT_START_TIMES for event_tp in "чз"],
        columns=["day", "time"],
    )

    for sheet, row_offset in zip(workbook.worksheets, [4, 3], strict=False):
        for col_idx, group in iter_students_groups(sheet, row_offset):
            df[[group]] = ""

            for event, idx in read_group_sheet(sheet, col_idx, row_offset + 1):
                df.at[idx, group] = event

    df["day"] = df["day"].str.replace("\n", "").ffill()
    df["day"] = df["day"].apply(lambda d: d.title())
    df["time"] = df["time"].apply(lambda x: tuple(x.split("_")))

    for cell in EMPTY_CELLS:
        df = df.replace(cell, None)

    return df


def normalize_teacher_name(name: str) -> str:
    surname, name, *_ = name.split()
    return f"{surname} {name}"


def iter_teachers_days_idx() -> Iterable[tuple[int, str, str]]:
    start = 2
    for day in DAYS:
        for t in EVENT_START_TIMES:
            yield start, day, t
            start += 1

        start += 1


def iter_through_teachers(sheet: Worksheet) -> Iterable[str]:
    for idx in count(8, 4):
        if name := sheet.cell(row=idx, column=1).value:
            yield idx, normalize_teacher_name(name)
        else:
            break


def parse_teacher_cells(cells: list[Cell]) -> Iterable[tuple[str, bool]]:
    match cells:
        case [
            Cell() as first,
            MergedCell(),
            MergedCell(),
            MergedCell(),
        ]:
            yield first.value, False
            yield first.value, True
        case [
            Cell(),
            Cell(),
            Cell() as second,
            MergedCell(),
        ]:
            yield second.value, True
        case [
            Cell() as first,
            MergedCell(),
            Cell() as second,
            MergedCell(),
        ]:
            yield first.value, False
            yield second.value, True


def iter_through_teacher_schedule(sheet: Worksheet, start_idx: int) -> Iterable[tuple[str, str]]:
    for idx, (cell_idx, _, _) in enumerate(iter_teachers_days_idx()):
        cells = [sheet.cell(row=start_idx + i, column=cell_idx) for i in range(4)]

        for event, is_denominator in parse_teacher_cells(cells):
            if not event:
                continue

            offset = idx * 2 + is_denominator
            yield offset, event


def get_teachers_schedule_df(url: str) -> pd.DataFrame:
    with fetch_schedule_io(url) as io:
        workbook = load_workbook(io)

    teachers = [teacher for _, teacher in iter_through_teachers(workbook.active)]

    df = pd.DataFrame.from_records(
        [
            {"day": day, "time": f"{time}_{event_tp}"} | dict.fromkeys(teachers, "")
            for day in DAYS
            for time in EVENT_START_TIMES
            for event_tp in "чз"
        ],
        columns=["day", "time", *teachers],
    )

    for idx, teacher in iter_through_teachers(workbook.active):
        for offset, event in iter_through_teacher_schedule(workbook.active, idx):
            df.at[offset, teacher] = event

    df["day"] = df["day"].str.replace("\n", "").ffill()
    df["day"] = df["day"].apply(lambda d: d.title())
    df["time"] = df["time"].apply(lambda x: tuple(x.split("_")))

    for cell in EMPTY_CELLS:
        df = df.replace(cell, None)

    return df


def get_group_name(subgroup: str) -> str:
    return re.sub(r"(\/|-)\d+.?$", "", subgroup)


def get_groups(df: pd.DataFrame) -> dict[str, list[str]]:
    groups = defaultdict(list)
    for subgroup in df.columns[2:]:
        group = re.sub(r"(\/|-)\d+.?$", "", subgroup)
        groups[group].append(subgroup)

    return groups


def all_same(items: Sequence[Any]) -> bool:
    return all(it == items[0] for it in items)


def normalize_subevents(events: list[str]) -> dict[str, Any] | None:
    if all_same(events):
        if not events[0]:
            return {"type": "empty"}

        return {"type": "single", "event": events[0]}

    return {"type": "multiple", "events": events}


def normalize_event(odd: list[str], even: list[str]) -> dict[str, Any] | None:
    if all(it is None for it in chain(odd, even)):  # empty event
        return None

    if odd == even:
        if all_same(odd):
            return {"type": "single", "event": odd[0]}

        return {
            "type": "vertical",
            "events": [{"type": "single", "event": val} if val else {"type": "empty"} for val in odd],
        }

    return {
        "type": "horizontal",
        "events": [
            normalize_subevents(odd),
            normalize_subevents(even),
        ],
    }


def get_grouped_schedule(df: pd.DataFrame, sub_entities: list[str]) -> Any:
    df = df[["day", "time", *sub_entities]]
    days: dict[str, dict[str, Any]] = defaultdict(lambda: defaultdict(lambda: ([], [])))

    for _, day, (time, week), *rest in df.itertuples():
        for event in rest:
            days[day][time][int(week != "ч")].append(event and event.strip())

    result = []
    for day, times in days.items():
        day_events = [
            {
                "time": f"{time} - {EVENT_END_TIMES[EVENT_START_TIMES.index(time)]}",
                "order": EVENT_START_TIMES.index(time) + 1,
                "event": event,
            }
            for time, events in times.items()
            if (event := normalize_event(*events))
        ]

        if day_events:
            result.append({"day": day, "events": day_events})

    return result


def get_teachers_schedule():
    df = get_teachers_schedule_df(TEACHERS_SCHEDULE_URL)
    teachers = [*df.columns[2:]]

    schedules = {
        teacher: {
            "teacher": teacher,
            "schedule": get_grouped_schedule(df, [teacher]),
        }
        for teacher in teachers
    }

    with (ROOT / "src" / "teachers.json").open("w") as f:
        json.dump(schedules, f, ensure_ascii=False, indent=4)


def get_students_schedule():
    df = get_schedule_df(STUDENTS_SCHEDULE_URL)
    groups = get_groups(df)

    schedules = {
        group: {
            "group": group,
            "subgroups": subgroups,
            "schedule": get_grouped_schedule(df, subgroups),
        }
        for group, subgroups in groups.items()
    }

    with (ROOT / "src" / "students.json").open("w") as f:
        json.dump(schedules, f, ensure_ascii=False, indent=4)


def main() -> None:
    get_students_schedule()
    get_teachers_schedule()


if __name__ == "__main__":
    main()
